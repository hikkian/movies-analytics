import os
import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from config import engine

os.makedirs("charts", exist_ok=True)
os.makedirs("exports", exist_ok=True)

def visualize_data():
    plots_info = []

    df1 = pd.read_sql("""
        SELECT g.name AS genre, COUNT(m.movie_id) AS movie_count
        FROM movies m
        JOIN genres g ON m.genre_id = g.genre_id
        GROUP BY g.name
    """, engine)
    df1.set_index('genre')['movie_count'].plot.pie(autopct='%1.1f%%', startangle=90)
    plt.title("Distribution of Movies by Genre")
    plt.ylabel("")
    plt.tight_layout()
    plt.savefig("charts/pie_genre_distribution.png")
    plt.close()
    plots_info.append({"type": "Pie Chart", "rows": len(df1), "desc": "Distribution of movies by genre"})

    df2 = pd.read_sql("""
        SELECT d.name AS director, AVG(r.rating) AS avg_rating
        FROM ratings r
        JOIN movies m ON r.movie_id = m.movie_id
        JOIN directors d ON m.director_id = d.director_id
        GROUP BY d.name
        HAVING COUNT(r.rating) >= 2
        ORDER BY avg_rating DESC
        LIMIT 10
    """, engine)
    plt.bar(df2['director'], df2['avg_rating'], color='steelblue')
    plt.title("Top 10 Directors by Average Rating (min 2 movies)")
    plt.xlabel("Director")
    plt.ylabel("Average Rating")
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    plt.savefig("charts/bar_director_avg_rating.png")
    plt.close()
    plots_info.append({"type": "Bar Chart", "rows": len(df2), "desc": "Top directors by avg rating"})

    df3 = pd.read_sql("""
        SELECT a.name AS actor, COUNT(ma.movie_id) AS movie_count
        FROM movie_actor ma
        JOIN actors a ON ma.actor_id = a.actor_id
        GROUP BY a.name
        ORDER BY movie_count DESC
        LIMIT 10
    """, engine)
    plt.barh(df3['actor'], df3['movie_count'], color='teal')
    plt.title("Top 10 Most Active Actors")
    plt.xlabel("Number of Movies")
    plt.ylabel("Actor")
    plt.tight_layout()
    plt.savefig("charts/hbar_actor_movie_count.png")
    plt.close()
    plots_info.append({"type": "Horizontal Bar", "rows": len(df3), "desc": "Most active actors"})

    df4 = pd.read_sql("""
        SELECT release_year, COUNT(movie_id) AS movie_count
        FROM movies
        WHERE release_year BETWEEN 1980 AND 2023
        GROUP BY release_year
        ORDER BY release_year
    """, engine)
    plt.plot(df4['release_year'], df4['movie_count'], marker='o', color='purple')
    plt.title("Movies Released Per Year (1980–2023)")
    plt.xlabel("Year")
    plt.ylabel("Number of Movies")
    plt.grid(True, alpha=0.5)
    plt.tight_layout()
    plt.savefig("charts/line_movies_per_year.png")
    plt.close()
    plots_info.append({"type": "Line Chart", "rows": len(df4), "desc": "Trend of movie releases"})

    df5 = pd.read_sql("SELECT duration_min FROM movies WHERE duration_min IS NOT NULL", engine)
    plt.hist(df5['duration_min'], bins=20, color='orange', edgecolor='black')
    plt.title("Distribution of Movie Durations")
    plt.xlabel("Duration (minutes)")
    plt.ylabel("Frequency")
    plt.grid(True, alpha=0.5)
    plt.tight_layout()
    plt.savefig("charts/hist_movie_duration.png")
    plt.close()
    plots_info.append({"type": "Histogram", "rows": len(df5), "desc": "Movie duration distribution"})

    df6 = pd.read_sql("""
        SELECT m.duration_min, r.rating
        FROM ratings r
        JOIN movies m ON r.movie_id = m.movie_id
        WHERE m.duration_min IS NOT NULL
    """, engine)
    plt.figure(figsize=(10, 6))
    plt.hexbin(df6['duration_min'], df6['rating'], gridsize=30, cmap='Greens', mincnt=1)
    plt.colorbar(label='Number of Movies')
    plt.title("Movie Rating vs. Duration (Density Heatmap)")
    plt.xlabel("Duration (minutes)")
    plt.ylabel("Rating")
    plt.grid(True, linestyle='--', alpha=0.5)
    plt.tight_layout()
    plt.savefig("charts/hexbin_rating_vs_duration.png")
    plt.close()

    for p in plots_info:
        print(f"✅ {p['type']}: {p['rows']} rows — {p['desc']}")
    return plots_info

def create_time_slider():
    df = pd.read_sql("""
        SELECT 
            m.release_year AS year,
            a.name AS actor,
            COUNT(ma.movie_id) AS movies,
            AVG(r.rating) AS avg_rating
        FROM movie_actor ma
        JOIN actors a ON ma.actor_id = a.actor_id
        JOIN movies m ON ma.movie_id = m.movie_id
        JOIN ratings r ON m.movie_id = r.movie_id
        WHERE m.release_year IS NOT NULL
        GROUP BY year, a.name
        HAVING COUNT(ma.movie_id) >= 1 AND AVG(r.rating) IS NOT NULL
        ORDER BY year
    """, engine)
    if df.empty:
        print("⚠️ No data for time slider")
        return
    df['year'] = df['year'].astype(int)
    fig = px.scatter(
        df, x="movies", y="avg_rating", animation_frame="year",
        size="movies", color="actor", hover_name="actor",
        range_x=[0, df['movies'].max() + 1],
        range_y=[df['avg_rating'].min() - 0.5, df['avg_rating'].max() + 0.5],
        title="Actor Performance Over Time"
    )
    print("✅ Plotly time-slider ready. Showing...")
    fig.show()

def export_to_excel(dataframes_dict, filename):
    path = os.path.join("exports", filename)
    with pd.ExcelWriter(path, engine='openpyxl') as w:
        for name, df in dataframes_dict.items():
            df.to_excel(w, sheet_name=name[:31], index=False)

    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions
        for col in ws.iter_cols():
            try:
                if col[1].value is not None and isinstance(col[1].value, (int, float)):
                    rng = f"{col[0].column_letter}2:{col[0].column_letter}{ws.max_row}"
                    rule = ColorScaleRule(
                        start_type="min", start_color="FFAA0000",
                        mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                        end_type="max", end_color="FF00AA00"
                    )
                    ws.conditional_formatting.add(rng, rule)
            except:
                continue
    wb.save(path)
    total = sum(len(df) for df in dataframes_dict.values())
    print(f"✅ Created file {filename}, {len(dataframes_dict)} sheets, {total} rows")

if __name__ == "__main__":
    visualize_data()
    create_time_slider()

    dfs = {
        "Genre_Count": pd.read_sql("SELECT g.name, COUNT(m.movie_id) FROM movies m JOIN genres g ON m.genre_id = g.genre_id GROUP BY g.name", engine),
        "Director_Ratings": pd.read_sql("SELECT d.name, AVG(r.rating) AS avg_rating FROM ratings r JOIN movies m ON r.movie_id = m.movie_id JOIN directors d ON m.director_id = d.director_id GROUP BY d.name", engine),
        "Actor_Count": pd.read_sql("SELECT a.name, COUNT(ma.movie_id) FROM movie_actor ma JOIN actors a ON ma.actor_id = a.actor_id GROUP BY a.name", engine)
    }
    export_to_excel(dfs, "movies_analysis_report.xlsx")